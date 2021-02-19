##########################################################
# Script principal do experimento
# Autor: Alesanco Andrade Azevedo
# Data: 2021-02-19
##########################################################

from sqlalchemy import Column, VARCHAR
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import create_engine
from openpyxl import load_workbook
from datetime import datetime
import os, sys, time
import pandas as pd
import numpy as np
from paramikos import *

engine  = create_engine('mysql://root:senha@127.0.0.1')
tables  = ['part','customer','supplier','dim_date','lineorder']
sf      = "sf1"
repeats = 30
cluster = " - 2 folhas - "
dir_experiments = "./experiments/"
hora = str(datetime.now())
stat1 = hora[0:10]+' - '+sf.upper()+cluster+'principal.csv'
stat2 = hora[0:10]+' - '+sf.upper()+cluster+'secundario.csv'
stat3 = hora[0:10]+' - '+sf.upper()+cluster+'zlixo.csv'

statistics1 = open(stat1,'a')
statistics2 = open(stat2,'a')
trash       = open(stat3,'a')
statistics1.write('date_time'+";"+'sf'+";"+'db'+";"+'part'+";"+'customer'+";"+'supplier'+";"+'dim_date'+";"+'lineorder'+";"+'total_time'+";"+'ram_mb'+";"+'disk_mb'+";"+'Q1.1'+";"+'Q1.2'+";"+'Q1.3'+";"+'Q2.1'+";"+'Q2.2'+";"+'Q2.3'+";"+'Q3.1'+";"+'Q3.2'+";"+'Q3.3'+";"+'Q3.4'+";"+'Q4.1'+";"+'Q4.2'+";"+'Q4.3'+";"+"\n")
statistics2.write('date_time'+";"+'sf'+";"+'db'+";"+'query_id'+";")
trash.write('date_time'+";"+'sf'+";"+'db'+";"+'query_id'+";")
for idx, a in enumerate(range(repeats)):
    statistics2.write('T'+str(idx+1)+";")
    trash.write('T'+str(idx+1)+";")
statistics2.write("mean;std;coef_v;\n")
trash.write("mean;std;coef_v;\n")
statistics1.close()
statistics2.close()
trash.close()

def remove_extrems(values):
    fator = 1.5
    q3, q1 = np.percentile(values, [75, 25])
    iqr = q3 - q1
    lowpass = q1 - (iqr * fator)
    highpass = q3 + (iqr * fator)
    return [v for v in values if v > lowpass and v < highpass]

def collect_resources(db):
    query = """ SELECT rowsegs.database_name, Memory_in_mb, Disk_in_mb FROM  (SELECT database_name, truncate(Round(Sum(memory_use / 1024 / 1024 ), 3),0) AS Memory_in_mb FROM 
    information_schema.table_statistics  GROUP BY database_name) rowsegs LEFT JOIN (SELECT database_name, truncate(Round(Sum(compressed_size) / 1024 / 1024 , 3),0) AS Disk_in_mb FROM
    information_schema.columnar_segments GROUP BY database_name) colsegs ON rowsegs.database_name = colsegs.database_name WHERE rowsegs.database_name = '%s';"""%(db)
    result = engine.execute(query)
    for res in result:
        try:
            return str(res[1]), str(res[2])
        except Exception:
            return 0,0

def create_db(experiment):
    try:
        sql = "%s%s"%(dir_experiments,experiment)
        file1 = open(sql,"r")
        query = ""
        for line in file1:
            query += line
        engine.execute(query)
        print "\nDatabase %s created."%(experiment[-6:-4])
        return 1
    except Exception:
        print "\nError in file %s."%(experiment)
        return 0
    
def load_db(db):
    hora = str(datetime.now())
    hora = hora[0:19]
    statistics1 = open(stat1,'a')
    statistics1.write(hora+";"+str(sf)+";"+str(db)+";")
    try:
        total_time = 0.0
        for table in tables:
            print "Insert data in table: %s" %(table),
            time1 = time.time()
            query = "LOAD DATA LOCAL INFILE './data/%s/%s.tbl' INTO TABLE %s.%s FIELDS TERMINATED BY '|' LINES TERMINATED BY '|\r\n'" %(sf, table, db, table)
            engine.execute(query)
            time2 = time.time()
            parcial_time = time2 - time1
            total_time = total_time + float(parcial_time)
            parcial_time = "%.2f" %(parcial_time)
            parcial_time = parcial_time.replace(".",",")
            print parcial_time
            statistics1.write(parcial_time+";")
        print "######"
        total_time = "%.2f" %(total_time)
        total_time = str(total_time)
        total_time = total_time.replace(".",",")
        print "Total time:" , total_time
        print "######"
        statistics1.write(total_time+";")
        lengths = collect_resources(db)
        statistics1.write(str(lengths[0])+";"+str(lengths[1])+";")
        statistics1.close()
        rebalance = "REBALANCE PARTITIONS ON %s;" %(db)
        engine.execute(rebalance)
        os.system("sudo sysctl -w vm.drop_caches=3")
        return 1
    except Exception:
        statistics1.write("error;")
        print "\nError in load of %s."%(str(db))
        statistics1.close()
        os.system("sudo sysctl -w vm.drop_caches=3")
        return 0

def load_db_unique(db):
    hora = str(datetime.now())
    hora = hora[0:19]
    statistics1 = open(stat1,'a')
    statistics1.write(hora+";"+str(sf)+";"+str(db)+";-;-;-;-;")
    print "Insert data in table: lineorder",
    try:
        time1 = time.time()
        query = "LOAD DATA LOCAL INFILE './data/%s/unique.csv' INTO TABLE %s.lineorder FIELDS TERMINATED BY '|' LINES TERMINATED BY '\n'" %(sf, db)
        engine.execute(query)
        time2 = time.time()
        parcial_time = time2 - time1
        parcial_time = "%.2f" %(parcial_time)
        parcial_time = parcial_time.replace(".",",")
        print parcial_time
        statistics1.write(parcial_time+";"+parcial_time+";")
        lengths = collect_resources(db)
        statistics1.write(str(lengths[0])+";"+str(lengths[1])+";")
        statistics1.close()
        os.system("sudo sysctl -w vm.drop_caches=3")
        return 1
    except Exception:
        statistics1.write("error;")
        print "\nError in load of %s."%(str(db))
        statistics1.close()
        os.system("sudo sysctl -w vm.drop_caches=3")
        return 0

def queries(db):
    query   = "USE %s"%(db)
    engine.execute(query)
    sheets  = []
    wb = ""
    if int(db[3:])<25:
        wb = load_workbook("queries.xlsx")
    else:
        wb = load_workbook("queries_unique.xlsx")
    names   = wb.sheetnames
    for name in names:
        sheets.append(name)
    #clear_caches(db)
    for sheet in sheets:
        ws = wb[sheet]
        for idx, line in enumerate(ws):
            if idx>0:
                query_id        = line[0].value
                query_content   = line[1].value
                statistics1     = open(stat1,'a')
                statistics2     = open(stat2,'a')
                try:
                    for k in range(10):
                        times = []
                        print query_id,"\t",
                        for a in range(repeats):
                            engine.execute("drop all from plancache;")
                            time1 = time.time()
                            engine.execute(query_content)
                            time2 = time.time()
                            time_query  = time2 - time1
                            time_query = "%.3f" %(time_query)
                            times.append(float(time_query))
                            if float(time_query)>100:
                                print "BREAK\t", time_query
                                break;
                        if len(times)!=1:
                            times = remove_extrems(times)
                        print times,
                        serie    =  pd.Series(times)
                        dp       = "%.3f" % (serie.std())
                        media    = "%.3f" % (serie.mean())
                        coef_var = "%.3f" % (serie.std() / serie.mean())
                        print "coef:", coef_var, "media:", media, "dp:", dp,
                        limiar = 0.1
                        if float(coef_var)<=limiar:
                            dp       = str(dp).replace(".",",")
                            media    = str(media).replace(".",",")
                            coef_var = str(coef_var).replace(".",",")
                            statistics2.write(str(datetime.now())+";"+str(sf)+";"+str(db)+";"+str(query_id)+";")
                            if len(times)>2: 
                                for t in times:
                                    t = str(t).replace(".",",")
                                    statistics2.write(str(t)+";")
                                if len(times)<repeats:
                                    ajust = repeats - len(times)
                                    for i in range(ajust):
                                        statistics2.write(";")
                                statistics2.write(media+";"+dp+";"+coef_var+";\n")
                            else:
                                for t in times:
                                    t = str(t).replace(".",",")
                                    statistics2.write(str(t)+";")
                                statistics2.write(";\n")
                            time_mean  = float("%.3f"%(sum(times)/len(times)))
                            time_mean = str(time_mean)
                            time_mean = time_mean.replace(".",",")
                            statistics1.write(time_mean+";")
                            print "OK"
                            break;
                        else:
                            print "nao"
                            dp       = str(dp).replace(".",",")
                            media    = str(media).replace(".",",")
                            coef_var = str(coef_var).replace(".",",")
                            trash    = open(stat3,'a')
                            trash.write(str(datetime.now())+";"+str(sf)+";"+str(db)+";"+str(query_id)+";")
                            if len(times)>2:
                                for t in times:
                                    t = str(t).replace(".",",")
                                    trash.write(str(t)+";")
                                if len(times)<repeats:
                                    ajust = repeats - len(times)
                                    for i in range(ajust):
                                        trash.write(";")
                                trash.write(media+";"+dp+";"+coef_var+";\n")
                            else:
                                for t in times:
                                    t = str(t).replace(".",",")
                                    trash.write(str(t)+";")
                                trash.write(";\n")
                            if k==9:
                                statistics1.write("-;")
                            trash.close()
                except Exception:
                    pass
                    statistics1.write("Error;")
                    statistics2.write(str(datetime.now())+";"+str(sf)+";"+str(db)+";"+str(query_id)+";")
                    for a in range(repeats+2):
                        statistics2.write("error;")
                    statistics2.write("\n")
                    print "Error in %s by query %s"%(str(db),str(query_id))
                statistics1.close()
                statistics2.close()
                cleaner = os.system("sudo sysctl -w vm.drop_caches=3")
    clear_caches(db)
    statistics1 = open(stat1,'a')
    statistics1.write("\n")
    statistics1.close()
    

def remove(db):
    sql = "DROP DATABASE IF EXISTS %s"%(db)
    engine.execute(sql)
    print "Database %s removed."%(db)
    os.system("sudo sysctl -w vm.drop_caches=3")
    print "##############################################"

def main():
    for i in range(1):
        for experiment in sorted(os.listdir(dir_experiments)):
            if "db" in experiment:
                db = "ssb%s"%(experiment[-6:-4])
                engine.execute("drop all from plancache;")
                sucess = create_db(experiment)
                if sucess:
                    if int(experiment[-6:-4])<25:
                        sucess = load_db(db)
                    else:
                        sucess = load_db_unique(db)
                    if sucess:
                        queries(db)
                remove(db)
                statistics1 = open(stat1,'a')
                statistics1.write("\n")
                statistics1.close()
                cleaner = os.system("sudo sysctl -w vm.drop_caches=3")
if __name__ == "__main__":
    main()
